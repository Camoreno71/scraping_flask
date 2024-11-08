from flask import Flask, jsonify
from playwright.sync_api import sync_playwright
from openpyxl import load_workbook, Workbook

app = Flask(__name__)

@app.route("/")
def scrape_page():

    # Ruta al archivo Excel
    archivo = 'CATALOGO JUAN DIEGO MACHADO.xlsx'
    # Cargar el archivo Excel
    wb = load_workbook(archivo)
    # Seleccionar la hoja activa (puedes también seleccionar por nombre con wb['Hoja1'])
    hoja = wb.active

    buscador = ""
    img_urls = []

    # Crear un nuevo libro de trabajo
    wbNewFile = Workbook()

    # Seleccionar la hoja activa
    hojaNew = wbNewFile.active
    aux = 1

    with sync_playwright() as p:
        # Iniciar el navegador y abrir una nueva página
        browser = p.chromium.launch(headless=False)  # Puedes poner headless=True para que no se vea el navegador
        page = browser.new_page()

        # Navegar a la URL deseada
        page.goto("https://elbaratillo.co/")

        print("Inicia Proceso------------------------")

        # Recorrer las filas de la hoja
        for fila in hoja.iter_rows(min_row=2, values_only=True):  # min_row=2 para saltarse la cabecera
            # 'fila' es una tupla con los valores de cada celda de la fila
            nombre_producto = fila[0]
            #print("producto: ", fila[0])  # Imprime todos los valores de la fila

            buscador = ""
            img_urls = []

            # Esperar a que el campo de entrada esté disponible
            page.wait_for_selector('input#dgwt-wcas-search-input-2')
            page.fill('input#dgwt-wcas-search-input-2', nombre_producto)  # Cambia el selector por el adecuado para tu input
            try:
                page.wait_for_selector('.dgwt-wcas-suggestion.dgwt-wcas-suggestion-product', timeout=2000)
                buscador = page.query_selector('.dgwt-wcas-suggestion.dgwt-wcas-suggestion-product')
            except Exception as e:
                print("No se encontro el producto")

            print("buscador: ", buscador)
            if buscador:
                page_product = buscador.get_attribute('href')
                print("url pagina", page_product)
                page.goto(page_product)
                page.wait_for_selector('li[aria-hidden="false"] img.attachment-woocommerce_gallery_thumbnail.size-woocommerce_gallery_thumbnail')
                li_elements = page.query_selector_all('li[aria-hidden="false"]')
                
                for li in li_elements:
                    img_element = li.query_selector('img.attachment-woocommerce_gallery_thumbnail.size-woocommerce_gallery_thumbnail')
                    if img_element:
                        lisrcset = img_element.get_attribute('srcset').split(", ")
                        img_url = lisrcset[-1].split(" ")
                        img_urls.append(img_url[0])

                #print("imgs product", img_urls)

                hojaNew.cell(row=aux, column=1, value=nombre_producto)

                for col_num, valor in enumerate(img_urls, start=2):  # Comenzamos en la columna 2
                    hojaNew.cell(row=aux, column=col_num, value=valor)
                
                aux += 1

        # Guardar el archivo Excel
        archivoFinal = "AgrupacionCatalogo.xlsx"
        wbNewFile.save(archivoFinal)
            
            

        # img_element = page.query_selector('img.zoomImg')
        # if img_element:
        #     img_url = img_element.get_attribute('src')
        # print(img_url)
        # page.wait_for_selector('li[aria-hidden="false"] img.attachment-woocommerce_gallery_thumbnail.size-woocommerce_gallery_thumbnail')
        # li_elements = page.query_selector_all('li[aria-hidden="false"]')
        # img_urls = []
        # for li in li_elements:
        #     img_element = li.query_selector('img.attachment-woocommerce_gallery_thumbnail.size-woocommerce_gallery_thumbnail')
            
        #     if img_element:
        #         img_url = img_element.get_attribute('src')
        #         img_urls.append(img_url)
        # print(img_urls)


        # print(p_element)
        # text = p_element.inner_text()
        # input_selector = 'input.hfe-search-form__input[type="search"]'
        # print(text)

        # page.fill(input_selector, 'texto de búsqueda')

        # button = page.query_selector('a.elementor-button.elementor-button-link.elementor-size-lg')
        # if button:
        #     button.click() 
        #     page.wait_for_load_state('networkidle')
        #     h2_element = page.query_selector('h2.widget-title')
        #     h2_text = h2_element.text_content()
        #     print(h2_text)

    return f"<p>hola mundo</p>"

if __name__ == "__main__":
    app.run(debug=True,port=5003)